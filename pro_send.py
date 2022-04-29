import openpyxl
from pathlib import Path
import pandas as pd
import smtplib, ssl
port = 465
sender_email = "vietnamhome.info@gmail.com"
message ="""Subject: Vietnam Home - Giới thiệu căn hộ cho người nước ngoài thuê
Chào anh/chị {name},

Em là Hà bên công ty Vietnam Home, bên em chuyên giới thiệu căn hộ, nhà ở cho người nước ngoài sống tại Việt Nam.

Anh/chị có thể tham khảo thêm thông tin về bên em tại Webside: https://hanoivietnamhome.com/

Được biết anh/chị hiện có căn hộ tại Roman Plaza.

Nếu anh/chị hiện có nhu cầu cho thuê anh (chị ) có thể cho em xin 1 số thông tin về căn hộ, địa chỉ căn hộ cho thuê để bên em có thể giới thiệu khách nhé ạ.

Anh/chị có thể để lại thông tin qua mail. Hoặc có thể liên hệ trực tiếp đến em qua số điện thoại: 0962526693 (SMS, Zalo, Viber)

Em rất xin lỗi nếu đã làm phiền.

Em xin chân thành cảm ơn ạ.

Em Thu Hà"""

password = input("Type your pasword and press enter: ")
context = ssl.create_default_context()
with smtplib.SMTP_SSL("smtp.gmail.com", port, context = context) as server:
    server.login(sender_email, password)
 
    data = pd.ExcelFile("Roman Plaza.xlsx")
    df = data.parse("Sheet1")
    ps = openpyxl.load_workbook('Roman Plaza.xlsx')
    sheet = ps['Sheet1']
    print(sheet.max_row)
    for row in range(2, sheet.max_row + 1):
        name= sheet['F' + str(row)].value
        email = sheet['K'+ str(row)].value
        server.sendmail(sender_email, email, message.format(name=name).encode("utf-8"))